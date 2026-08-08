"""
Portfolio Sector Weightage Checker
Uses Yahoo Finance (yfinance) to fetch sector info for NSE stocks.

Usage:
    pip install yfinance openpyxl
    python sector_weightage.py
"""

import yfinance as yf
from collections import defaultdict

# ── Portfolio: ticker → weightage % ──────────────────────────────────────────
PORTFOLIO = {
    "ONGC.NS":       3.17,
    "BANKINDIA.NS":  3.51,
    "BHARATFORG.NS": 3.42,
    "BSE.NS":        3.17,
    "COALINDIA.NS":  3.17,
    "CUMMINSIND.NS": 3.17,
    "FEDERALBNK.NS": 3.77,
    "GESHIP.NS":     3.17,
    "HINDALCO.NS":   4.54,
    "INDIANB.NS":    3.89,
    "KARURVYSYA.NS": 3.17,
    "MAHABANK.NS":   3.25,
    "MCX.NS":        3.97,
    "NATIONALUM.NS": 6.33,
    "NAVINFLUOR.NS": 3.17,
    "SBIN.NS":       4.27,
    "SHRIRAMFIN.NS": 3.81,
    "TORNTPOWER.NS": 3.17,
    "UNIONBANK.NS":  3.80,
    "VEDL.NS":       4.71,
}

# ── Fetch sector info from Yahoo Finance ─────────────────────────────────────
print("\nFetching sector data from Yahoo Finance...\n")
print(f"{'Ticker':<18} {'Sector':<30} {'WGT%':>6}")
print("─" * 58)

ticker_sector = {}
failed = []

for ticker, wgt in PORTFOLIO.items():
    try:
        info = yf.Ticker(ticker).info
        sector = info.get("sector") or "Unknown"
        ticker_sector[ticker] = sector
        print(f"{ticker:<18} {sector:<30} {wgt:>5.2f}%")
    except Exception as e:
        ticker_sector[ticker] = "Unknown"
        failed.append(ticker)
        print(f"{ticker:<18} {'[ERROR - Unknown]':<30} {wgt:>5.2f}%")

# ── Aggregate by sector ───────────────────────────────────────────────────────
sector_wgt = defaultdict(float)
for ticker, wgt in PORTFOLIO.items():
    sector = ticker_sector[ticker]
    sector_wgt[sector] += wgt

sector_wgt = dict(sorted(sector_wgt.items(), key=lambda x: -x[1]))

# ── Print sector summary ──────────────────────────────────────────────────────
print("\n" + "═" * 45)
print(f"  {'SECTOR SUMMARY':^41}  ")
print("═" * 45)
print(f"{'Sector':<30} {'WGT%':>8}")
print("─" * 45)
for sector, wgt in sector_wgt.items():
    bar = "█" * int(wgt / 1.5)
    print(f"{sector:<30} {wgt:>6.2f}%  {bar}")
print("─" * 45)
print(f"{'TOTAL':<30} {sum(sector_wgt.values()):>6.2f}%")
print("═" * 45)

if failed:
    print(f"\n⚠  Could not fetch data for: {', '.join(failed)}")
    print("   These are counted under 'Unknown' sector.")

# ── Export to Excel ───────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sectoral Weightage"

    header_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    alt_fill    = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")
    white_fill  = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
    total_fill  = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
    total_font  = Font(name="Arial", bold=True, size=10, color="1F4E79")
    row_font    = Font(name="Arial", size=10)
    title_font  = Font(name="Arial", bold=True, size=13, color="1F4E79")
    thin        = Side(style="thin", color="BFBFBF")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:B1")
    ws["A1"] = "Portfolio — Sectoral Weightage"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.row_dimensions[2].height = 6

    for col, header in [("A3", "Sector"), ("B3", "WGT%")]:
        ws[col] = header
        ws[col].font = header_font
        ws[col].fill = header_fill
        ws[col].alignment = Alignment(horizontal="center", vertical="center")
        ws[col].border = border
    ws.row_dimensions[3].height = 20

    for i, (sector, wgt) in enumerate(sector_wgt.items(), start=4):
        fill = alt_fill if i % 2 == 0 else white_fill
        ws[f"A{i}"] = sector
        ws[f"B{i}"] = round(wgt, 2)
        for col in [f"A{i}", f"B{i}"]:
            ws[col].font = row_font
            ws[col].fill = fill
            ws[col].border = border
        ws[f"A{i}"].alignment = Alignment(horizontal="left", vertical="center")
        ws[f"B{i}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"B{i}"].number_format = '0.00"%"'
        ws.row_dimensions[i].height = 18

    tr = 4 + len(sector_wgt)
    ws[f"A{tr}"] = "TOTAL"
    ws[f"B{tr}"] = f"=SUM(B4:B{tr - 1})"
    for col in [f"A{tr}", f"B{tr}"]:
        ws[col].font = total_font
        ws[col].fill = total_fill
        ws[col].alignment = Alignment(horizontal="center", vertical="center")
        ws[col].border = border
    ws[f"B{tr}"].number_format = '0.00"%"'
    ws.row_dimensions[tr].height = 20

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14

    out_file = "sectoral_weightage.xlsx"
    wb.save(out_file)
    print(f"\n✅  Excel saved → {out_file}")

except ImportError:
    print("\n⚠  openpyxl not installed — skipping Excel export.")
    print("   Run: pip install openpyxl")